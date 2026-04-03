import pytest
import tempfile
import os
from pathlib import Path

import openpyxl


@pytest.fixture
def single_sheet_excel(tmp_path: Path) -> Path:
    file_path = tmp_path / "single_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "age", "city"])
    ws.append(["John", 30, "New York"])
    ws.append(["Jane", 25, "Boston"])
    wb.save(file_path)
    return file_path


@pytest.fixture
def multi_sheet_excel(tmp_path: Path) -> Path:
    file_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Users"
    ws1.append(["id", "name"])
    ws1.append([1, "Alice"])
    ws1.append([2, "Bob"])

    ws2 = wb.create_sheet("Orders")
    ws2.append(["order_id", "user_id", "amount"])
    ws2.append([101, 1, 99.99])
    ws2.append([102, 2, 149.50])
    ws2.append([103, 1, 25.00])

    wb.save(file_path)
    return file_path


@pytest.fixture
def empty_excel(tmp_path: Path) -> Path:
    file_path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    wb.save(file_path)
    return file_path


@pytest.fixture
def single_sheet_excel_bytes(single_sheet_excel: Path) -> bytes:
    with open(single_sheet_excel, "rb") as f:
        return f.read()
