"""LangChain Tool tests for ExcelExportTool."""
import io
import json
import os
import tempfile
from unittest.mock import MagicMock

import pytest

from src.infrastructure.excel_export.excel_export_tool import ExcelExportTool


@pytest.fixture
def tool():
    return ExcelExportTool()


@pytest.fixture
def valid_input_json():
    return json.dumps({
        "columns": ["Name", "Score"],
        "rows": [["Alice", 90], ["Bob", 85]],
        "filename": "result.xlsx",
        "sheet_name": "Report",
    })


class TestExcelExportTool:
    def test_tool_name_is_excel_export(self, tool):
        assert tool.name == "excel_export"

    def test_tool_has_description(self, tool):
        assert len(tool.description) > 10

    def test_tool_has_args_schema(self, tool):
        assert tool.args_schema is not None

    def test_run_creates_xlsx_file_and_returns_path(self, tool, tmp_path):
        result = tool._run(
            columns=["Name", "Score"],
            rows=[["Alice", 90], ["Bob", 85]],
            filename="result.xlsx",
            sheet_name="Report",
            output_dir=str(tmp_path),
        )
        assert result.endswith(".xlsx")
        assert os.path.exists(result)

    def test_run_file_contains_correct_data(self, tool, tmp_path):
        import openpyxl

        path = tool._run(
            columns=["Name", "Score"],
            rows=[["Alice", 90], ["Bob", 85]],
            filename="result.xlsx",
            sheet_name="Data",
            output_dir=str(tmp_path),
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Data"]
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"

    def test_run_uses_tmp_dir_by_default(self, tool):
        result = tool._run(
            columns=["A"],
            rows=[[1]],
            filename="test.xlsx",
        )
        assert os.path.exists(result)
        os.remove(result)  # cleanup

    def test_run_multiple_sheets_via_extra_sheets(self, tool, tmp_path):
        import openpyxl

        extra_sheets = [
            {"sheet_name": "Sheet2", "columns": ["X"], "rows": [[42]]}
        ]
        path = tool._run(
            columns=["A"],
            rows=[[1]],
            filename="multi.xlsx",
            sheet_name="Sheet1",
            output_dir=str(tmp_path),
            extra_sheets=extra_sheets,
        )
        wb = openpyxl.load_workbook(path)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames

    def test_run_returns_error_message_on_failure(self, tool, tmp_path):
        from unittest.mock import patch
        with patch(
            "src.infrastructure.excel_export.excel_export_tool.PandasExcelExporter.export",
            side_effect=RuntimeError("export failed"),
        ):
            result = tool._run(
                columns=["A"],
                rows=[[1]],
                filename="fail.xlsx",
                output_dir=str(tmp_path),
            )
        assert "ERROR" in result or "error" in result.lower()

    async def test_arun_delegates_to_run(self, tool, tmp_path):
        result = await tool._arun(
            columns=["Name"],
            rows=[["Alice"]],
            filename="async.xlsx",
            output_dir=str(tmp_path),
        )
        assert result.endswith(".xlsx")
        assert os.path.exists(result)
