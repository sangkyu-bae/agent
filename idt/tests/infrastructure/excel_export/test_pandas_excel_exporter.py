"""Infrastructure tests for PandasExcelExporter.

Uses mocking for pandas/openpyxl. Also tests real export for integration.
"""
import io
from unittest.mock import MagicMock, patch

import pytest

from src.domain.excel_export.schemas import ExcelExportRequest, ExcelSheetData
from src.infrastructure.excel_export.pandas_excel_exporter import PandasExcelExporter


@pytest.fixture
def exporter():
    return PandasExcelExporter()


@pytest.fixture
def single_sheet_request():
    return ExcelExportRequest(
        filename="report.xlsx",
        sheets=[
            ExcelSheetData(
                sheet_name="Sales",
                columns=["Name", "Amount"],
                rows=[["Alice", 1000], ["Bob", 2000]],
            )
        ],
        request_id="req-001",
        user_id="user-1",
    )


@pytest.fixture
def multi_sheet_request():
    return ExcelExportRequest(
        filename="report.xlsx",
        sheets=[
            ExcelSheetData(
                sheet_name="Sheet1",
                columns=["A", "B"],
                rows=[[1, 2], [3, 4]],
            ),
            ExcelSheetData(
                sheet_name="Sheet2",
                columns=["X", "Y"],
                rows=[["a", "b"]],
            ),
        ],
        request_id="req-002",
        user_id="user-1",
    )


class TestPandasExcelExporter:
    def test_get_exporter_name_returns_pandas_openpyxl(self, exporter):
        assert exporter.get_exporter_name() == "pandas+openpyxl"

    def test_export_returns_excel_export_result(self, exporter, single_sheet_request):
        result = exporter.export(single_sheet_request)

        assert result.filename == "report.xlsx"
        assert result.user_id == "user-1"
        assert result.request_id == "req-001"
        assert result.sheet_count == 1
        assert result.exporter_used == "pandas+openpyxl"
        assert result.size_bytes > 0
        assert len(result.excel_bytes) > 0

    def test_export_produces_valid_xlsx_bytes(self, exporter, single_sheet_request):
        """생성된 bytes가 실제 xlsx 파일임을 검증 (PK 매직바이트)."""
        import openpyxl

        result = exporter.export(single_sheet_request)
        wb = openpyxl.load_workbook(io.BytesIO(result.excel_bytes))
        assert "Sales" in wb.sheetnames

    def test_export_writes_correct_data_to_cells(self, exporter, single_sheet_request):
        import openpyxl

        result = exporter.export(single_sheet_request)
        wb = openpyxl.load_workbook(io.BytesIO(result.excel_bytes))
        ws = wb["Sales"]

        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Amount"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 1000
        assert ws.cell(3, 1).value == "Bob"
        assert ws.cell(3, 2).value == 2000

    def test_export_multi_sheet_creates_all_sheets(self, exporter, multi_sheet_request):
        import openpyxl

        result = exporter.export(multi_sheet_request)
        assert result.sheet_count == 2

        wb = openpyxl.load_workbook(io.BytesIO(result.excel_bytes))
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames

    def test_export_empty_rows_creates_header_only_sheet(self, exporter):
        import openpyxl

        request = ExcelExportRequest(
            filename="empty.xlsx",
            sheets=[ExcelSheetData(columns=["Col1", "Col2"], rows=[])],
            request_id="req-001",
            user_id="user-1",
        )
        result = exporter.export(request)
        wb = openpyxl.load_workbook(io.BytesIO(result.excel_bytes))
        ws = wb.active
        assert ws.cell(1, 1).value == "Col1"
        assert ws.cell(2, 1).value is None

    def test_export_size_bytes_matches_actual_bytes(self, exporter, single_sheet_request):
        result = exporter.export(single_sheet_request)
        assert result.size_bytes == len(result.excel_bytes)

    def test_export_raises_runtime_error_on_pandas_failure(self, exporter):
        request = ExcelExportRequest(
            filename="fail.xlsx",
            sheets=[ExcelSheetData(columns=["A"], rows=[])],
            request_id="req-001",
            user_id="user-1",
        )
        with patch(
            "src.infrastructure.excel_export.pandas_excel_exporter.pd.DataFrame",
            side_effect=Exception("pandas crash"),
        ):
            with pytest.raises(RuntimeError, match="Excel 생성 중 오류가 발생했습니다"):
                exporter.export(request)
